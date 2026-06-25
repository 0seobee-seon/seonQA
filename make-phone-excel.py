from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "사내전화번호"

headers = ["이름", "직함", "내선번호", "부서", "근무지", "외선전화", "비고"]
data = [
    ["김영섭", "차장", "505", "총무팀", "본사 3F", "", ""],
    ["박지은", "부장", "500", "총무팀", "본사 3F", "", ""],
    ["박근동", "부사장", "535", "임원", "본사 1~2F", "", ""],
    ["홍상희", "전무", "9*2-120", "진단팀", "수동사무실", "255-4710", ""],
    ["김태중", "대표이사", "9*2-300", "건설본부", "수동사무실", "222-4590", ""],
]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=11)
data_font = Font(name="맑은 고딕", size=10)
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for row_idx, row in enumerate(data, 2):
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row_idx].height = 18

for r in range(len(data) + 2, len(data) + 22):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.font = data_font
        cell.alignment = center

col_widths = [12, 10, 12, 14, 14, 14, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

out = r"C:\Users\LG\Desktop\AX 코드 학습\챗봇 만들기\database\사내전화번호_양식.xlsx"
wb.save(out)
print("saved:", out)
